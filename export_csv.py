import csv
import sqlite3
from openpyxl import Workbook


conn = sqlite3.connect("un.db")
cursor = conn.cursor()

country_headings = ["id", "name"]


def query_general_info():
    general_headings = [
        "region",
        "membership_date",
        "population",
        "population_density",
        "surface_area",
        "sex_ratio",
        "capital_city",
        "capital_population",
        "currency",
        "exchange_rate",
    ]

    general_sql_headings = [
        f"Countries.'{heading}'" for heading in country_headings
    ] + [f"GeneralInfo.'{heading}'" for heading in general_headings]

    general_sql_headings_string = ", ".join(general_sql_headings)

    general_sql_stmt = """
    SELECT %s FROM Countries JOIN GeneralInfo ON Countries.id=GeneralInfo.country_id
    """ % (
        general_sql_headings_string
    )

    results = cursor.execute(general_sql_stmt).fetchall()
    return results


for row in query_general_info():
    print(row)


def query_econ_indicators():
    econ_headings = [
        "year",
        "GDP",
        "GDP growth",
        "GDP per capita",
        "Agriculture",
        "Industry",
        "Services",
        "Agriculture employment",
        "Industry employment",
        "Services employment",
        "Unemployment rate",
        "Labour participation (females)",
        "Labour participation (males)",
        "CPI",
        "Agricultural index",
        "Industrial index",
        "Exports",
        "Imports",
        "Trade balance",
        "Current account",
    ]

    econ_sql_headings = [f"Countries.'{heading}'" for heading in country_headings] + [
        f"EconIndicator.'{heading}'" for heading in econ_headings
    ]

    econ_sql_headings_string = ", ".join(econ_sql_headings)

    econ_sql_stmt = """
    SELECT %s FROM Countries JOIN EconIndicator ON Countries.id=EconIndicator.country_id
    """ % (
        econ_sql_headings_string
    )

    results = cursor.execute(econ_sql_stmt).fetchall()
    return results


def query_social_indicators():
    social_headings = [
        "year",
        "Population growth",
        "Urban population",
        "Urban population growth",
        "Fertility rate",
        "Life expectancy (females)",
        "Life expectancy (males)",
        "Population distribution (children)",
        "Population distribution (elders)",
        "Migrant",
        "Migrant ratio",
        "Refugees",
        "Infant mortality",
        "Health expenditure",
        "Physicians ratio",
        "Education expenditure",
        "Primary enroll ratio (females)",
        "Primary enroll ratio (males)",
        "Secondary enroll ratio (females)",
        "Secondary enroll ratio (males)",
        "Tertiary enroll ratio (females)",
        "Tertiary enroll ratio (males)",
        "Homicide rate",
        "Women in parliaments",
    ]

    social_sql_headings = [f"Countries.'{heading}'" for heading in country_headings] + [
        f"SocialIndicator.'{heading}'" for heading in social_headings
    ]

    social_sql_headings_string = ", ".join(social_sql_headings)

    social_sql_stmt = """
    SELECT %s FROM Countries JOIN SocialIndicator ON Countries.id=SocialIndicator.country_id
    """ % (
        social_sql_headings_string
    )

    results = cursor.execute(social_sql_stmt).fetchall()
    return results


def query_env_indicators():
    env_headings = [
        "year",
        "Internet usage",
        "Research expenditure",
        "Threatened species",
        "Forested area",
        "CO2 emission",
        "CO2 emission (per capita)",
        "Energy production",
        "Energy supply",
        "Tourists",
        "Important sites",
        "Drinking water (urban)",
        "Drinking water (rural)",
        "Sanitation (urban)",
        "Sanitation (rural)",
        "Assist disbursed",
        "Assist received",
    ]

    env_sql_headings = [f"Countries.'{heading}'" for heading in country_headings] + [
        f"EnvIndicator.'{heading}'" for heading in env_headings
    ]

    env_sql_headings_string = ", ".join(env_sql_headings)

    env_sql_stmt = """
    SELECT %s FROM Countries JOIN EnvIndicator ON Countries.id=EnvIndicator.country_id
    """ % (
        env_sql_headings_string
    )

    results = cursor.execute(env_sql_stmt).fetchall()
    return results


HEADING_COUNTRY = ["id", "Country Name"]

HEADING_GENERAL_INFO = [
    "Region",
    "UN membership date",
    "Population (000, 2018)",
    "Pop. density (per km2, 2018)",
    "Surface area (km2)",
    "Sex ratio (m per 100 f)",
    "Capital city",
    "Capital city pop. (000, 2018)",
    "National currency",
    "Exchange rate (per US$)",
]

HEADING_ECON_INDICATORS = [
    "Year",
    "GDP",
    "GDP grow rate (%)",
    "GDP per capita",
    "Economy: Agriculture (%)",
    "Economy: Industry (%)",
    "Economy: Services (%)",
    "Employment in agriculture (%)",
    "Employment in industry (%)",
    "Employment in services (%)",
    "Unemployment (%)",
    "Labour force participation (females) (%)",
    "Labour force participation (males) (%)",
    "CPI",
    "Agriculture production index",
    "International trade: exports",
    "International trade: imports",
    "Interantional trade: balance",
    "Current account",
]

HEADING_SOCIAL_INDICATORS = [
    "Year",
    "Population growth rate (%)",
    "Urban population (%)",
    "Urabn population growth rate (%)",
    "Fertility rate",
    "Life expectancy (females)",
    "Life expectancy (males)",
    "Population age distribution (0-14) (%)",
    "Population age distribution (60+) (%)",
    "International migrant stock",
    "International migrant stock (%)",
    "Refugees",
    "Infant mortality rate (%)",
    "Health: Current expenditure (%)",
    "Health: Physicians",
    "Education: Government expenditure (%)",
    "Education: Primary enrol. ratio (females)",
    "Education: Primary enrol. ratio (males)",
    "Education: Seconday enrol. ratio (females)",
    "Education: Seconday enrol. ratio (males)",
    "Education: Tertiary enrol. ratio (females)",
    "Education: Tertiary enrol. ratio (males)",
    "Intentional homicide rate",
    "Seats by women in National Parliaments",
]

HEADING_ENV_INDICATORS = [
    "Individuals using Internet",
    "Research & Development expenditure (%)",
    "Threatened species",
    "Forested area (%)",
    "CO2 emission",
    "CO2 emission (per capital)",
    "Enery production",
    "Energy supply",
    "Tourist arrivals",
    "Important sites for biodiversity protected",
    "Pop. with drinking water (urban) (%)",
    "Pop. with drinking water (rural) (%)",
    "Pop. with sanitation (urban) (%)",
    "Pop. with sanitation (rural) (%)",
    "Net assist disbursed (%)",
    "Net assist received (%)",
]

wb = Workbook()
wg = wb.create_sheet("General Information")
wg.append(HEADING_COUNTRY + HEADING_GENERAL_INFO)

for row in query_general_info():
    wg.append(row)

we = wb.create_sheet("Economic Indicators")
we.append(HEADING_COUNTRY + HEADING_ECON_INDICATORS)

for row in query_econ_indicators():
    we.append(row)

ws = wb.create_sheet("Social Indicators")
ws.append(HEADING_COUNTRY + HEADING_SOCIAL_INDICATORS)

for row in query_social_indicators():
    ws.append(row)

wv = wb.create_sheet("Environment Indicators")
wv.append(HEADING_COUNTRY + HEADING_ENV_INDICATORS)

for row in query_env_indicators():
    wv.append(row)

wb.save("un_data.xlsx")
