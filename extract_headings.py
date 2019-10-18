import requests
from bs4 import BeautifulSoup
import sqlite3
from openpyxl import Workbook

conn = sqlite3.connect("/mnt/s/Downloads/un.sqlite")
cursor = conn.cursor()

countries = cursor.execute("SELECT * FROM Countries").fetchall()

wb = Workbook()
ws_general = wb.create_sheet("General")
ws_social = wb.create_sheet("Social")
ws_econ = wb.create_sheet("Econ")
ws_env = wb.create_sheet("Env")
workbooks = {
    "General Information": ws_general,
    "Economic indicators": ws_econ,
    "Social indicators": ws_social,
    "Environment and infrastructure indicators": ws_env,
}


BASE_URL = "http://data.un.org/"
for country in countries:
    url = BASE_URL + country[2]
    country_name = country[1]
    print("Retriving country: ", country)

    r = requests.get(url)
    if r.status_code == 200:
        html = r.text
    else:
        print("Error encountered in:", country_name)
        print("Status code: ", r.status_code)
        continue

    parsed_html = BeautifulSoup(html, "html.parser")
    table_names = [table.summary.text for table in parsed_html.find_all("details")]
    tables = parsed_html.find_all("details")
    table_dict = dict()
    for table in tables:
        if table.summary.text == "General Information":
            table_dict["General Information"] = table
        elif table.summary.text == "Social indicators":
            table_dict["Social indicators"] = table
        elif table.summary.text == "Economic indicators":
            table_dict["Economic indicators"] = table
        elif table.summary.text == "Environment and infrastructure indicators":
            table_dict["Environment and infrastructure indicators"] = table

    for name, table in table_dict.items():
        rows = table.find_all("tr")
        fields = [country_name]
        for row in rows:
            cells = row.find_all("td")
            field = cells[0].text.replace("\xa0", "")
            fields.append(field)
        workbooks[name].append(fields)
        print(f"{country_name}: Table {name} updated.")

wb.save(filename="/mnt/s/Downloads/headings.xlsx")

